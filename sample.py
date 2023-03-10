from openpyxl import load_workbook

# Load the workbook using a relative path
workbook = load_workbook(filename='data/data.xlsx')

# Select the worksheet
worksheet = workbook.active

# Access cells in the worksheet
cell_value = worksheet['A1'].value
print(cell_value)

# Access rows in the worksheet
for row in worksheet.iter_rows(min_row=2, values_only=True):
    print(row)
