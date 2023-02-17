import os
import sys
from openpyxl import load_workbook

if len(sys.argv) < 2:
    print("Please provide the path to the Excel file as a command line argument.")
    sys.exit(1)

excel_file_path = os.environ['EXCEL_FILE_PATH']

workbook = load_workbook(filename=excel_file_path)
worksheet = workbook.active

for row in worksheet.iter_rows(min_row=2, values_only=True):
    # Do something with the data in each row
    print(row)

workbook.close()
